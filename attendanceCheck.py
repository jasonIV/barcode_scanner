from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()
ws = wb.active

#ws.append(['barCodeID', 'Day1', 'Day2', 'Day3', 'Day4', 'Day5', 'Day6', 'Day7', 'Day8', 'Counter'])

# x = 7
header = ['BarCodeID']

header.append('19110011')
header.append('19110028')
header.append('19110035')
header.append('19110042')
header.append('19110059')
ws.append(header)


for column in range(1,7):
    ws.column_dimensions[chr(64+column)].width = len(header[column-1])+20

for column in range(2,7):
    ws['{}26'.format(chr(64+column))] = '0'
    ws['{}27'.format(chr(64+column))] = '0'

wb1 = load_workbook(filename='attendanceCheck.xlsx', read_only=True)
ws1 = wb1.worksheets[0]
# list_with_values=[]
# for cell in ws1[1]:
#     list_with_values.append(cell.value)
# print(list_with_values)

for row in range(2, 26):
    ws['A{}'.format(row)] = 'Day{}'.format(row-1)
ws['A26'] = 'Count'
ws['A27'] = 'DateTime'
wb.save('attendanceCheck.xlsx')